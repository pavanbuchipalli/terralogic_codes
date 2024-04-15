import openpyxl
#to read the data from file
file = "bchdj" #path ofthe file
workbook  = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"] #name of the sheet is sheet1
# sheet = workbook.active   #if u only have one sheet
row = sheet.max_row
col = sheet.max_column
for i  in range(1,row+1):
    for j in range(1,col+1):
        print(sheet.cell(i,j).value)

#to write the similar data into the file

file = "bchdj" #path ofthe file
workbook  = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"] #name of the sheet is sheet1
# sheet = workbook.active   #if u only have one sheet
row = sheet.max_row
col = sheet.max_column
for i  in range(1,row+1):
    for j in range(1,col+1):
        sheet.cell(i,j).value="welcome"
workbook.save(file)
#to write multiple date into excel file


file = "bchdj" #path ofthe file
workbook  = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"] #name of the sheet is sheet1
# sheet = workbook.active   #if u only have one sheet
row = sheet.max_row
col = sheet.max_column

sheet.cell(1,1).value="welcome"
sheet.cell(1,2).valeu = 123

sheet.cell(2,1).value="welcome2"
sheet.cell(2,2).valeu = 1234

sheet.cell(3,1).value="welcome2"
sheet.cell(3,2).valeu = 12334
workbook.save(file)